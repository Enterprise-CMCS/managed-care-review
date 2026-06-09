#!/usr/bin/env python3
"""
Generate a DSNP-contracts spreadsheet.

Runs dsnp_contracts.sql against the local DB (loaded with prod data), resolves
each contract's MCR-XX-NNNN-PROGRAM name from statePrograms.json (the program
short-names are not stored in the database), and writes the .xlsx (default:
~/Documents/DSNP_Contracts_in_MC-Review.xlsx).

Usage (DATABASE_URL is read from the environment, e.g. .envrc / direnv):
    python3 generate_dsnp_spreadsheet.py                      # all DSNP contracts
    python3 generate_dsnp_spreadsheet.py 2025-11-04           # submitted on/after a date
    python3 generate_dsnp_spreadsheet.py 2025-11-04 2025-12-31  # within a date range
    python3 generate_dsnp_spreadsheet.py 2025-11-04 --by latest  # filter on latest submission
    python3 generate_dsnp_spreadsheet.py -o report.xlsx      # custom output path

Normally invoked via the ./generate-dsnp-report wrapper.
Requires: psql on PATH, openpyxl (pip install openpyxl).
"""
import argparse
import csv
import io
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
SQL_FILE = Path(__file__).resolve().parent / "dsnp_contracts.sql"
STATE_PROGRAMS = (
    REPO_ROOT / "packages/submissions/src/statePrograms/statePrograms.json"
)

HEADERS = [
    "Contract Name",
    "State",
    "contractID",
    "federalAuthorities",
    "rateMedicaidPopulations",
    "dsnpContract",
    "Initial Submission Date",
    "Latest Submission Date",
]


def load_program_lookup():
    """programID -> short name, exactly as the app's programNames() reads it."""
    data = json.loads(STATE_PROGRAMS.read_text())
    lookup = {}
    for state in data["states"]:
        for prog in state["programs"]:
            lookup[prog["id"]] = prog["name"]
    return lookup


def natural_key(s):
    """Approximates JS localeCompare(s, 'en', {numeric: true}) for sorting
    program names: case-insensitive, with digit runs compared numerically."""
    parts = re.split(r"(\d+)", s.lower())
    return [int(p) if p.isdigit() else p for p in parts]


def package_name(state_code, state_number, program_ids, lookup):
    """Port of packageName() in packages/submissions/src/contract/contractHelpers.ts."""
    pad = str(state_number).zfill(4)
    names = [lookup.get(pid, "Unknown Program") for pid in program_ids if pid]
    names.sort(key=natural_key)
    formatted = []
    for n in names:
        n = n.replace(" ", "-")           # \s -> -
        n = re.sub(r"[^a-zA-Z0-9+]", "", n)  # keep alphanumerics and +
        formatted.append(n.upper())
    return f"MCR-{state_code.upper()}-{pad}-{'-'.join(formatted)}"


def psql_url(database_url):
    """Strip Prisma-only query params (schema, connection_limit, ...) that
    libpq's URI parser rejects."""
    base, _, query = database_url.partition("?")
    if not query:
        return database_url
    keep = [kv for kv in query.split("&")
            if kv.split("=")[0] in ("sslmode", "host", "port", "user",
                                    "password", "dbname", "application_name")]
    return base + ("?" + "&".join(keep) if keep else "")


def run_query(database_url, from_date, to_date, basis):
    cmd = [
        "psql", psql_url(database_url),
        "-v", f"from_date={from_date or ''}",
        "-v", f"to_date={to_date or ''}",
        "-v", f"basis={basis}",
        "--csv",
        "-f", str(SQL_FILE),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        sys.exit(f"psql failed:\n{result.stderr}")
    return list(csv.DictReader(io.StringIO(result.stdout)))


def valid_date(s):
    try:
        datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        raise argparse.ArgumentTypeError(f"{s!r} is not a valid YYYY-MM-DD date")
    return s


def main():
    ap = argparse.ArgumentParser(
        description="Generate a spreadsheet of DSNP contracts. With no dates it "
                    "captures all DSNP contracts up to today; pass a start date "
                    "for that date onward, or start + end for a range. Use --by "
                    "to choose whether dates filter on each contract's initial or "
                    "latest submission date.")
    ap.add_argument("from_date", nargs="?", type=valid_date, metavar="FROM",
                    help="only contracts submitted on/after this date (YYYY-MM-DD)")
    ap.add_argument("to_date", nargs="?", type=valid_date, metavar="TO",
                    help="only contracts submitted on/before this date (YYYY-MM-DD)")
    ap.add_argument("--by", choices=("initial", "latest"), default="initial",
                    help="which submission date the FROM/TO window filters on "
                         "(default: initial)")
    ap.add_argument("-o", "--out",
                    default=os.path.expanduser("~/Documents/DSNP_Contracts_in_MC-Review.xlsx"),
                    help="output .xlsx path (default: ~/Documents/DSNP_Contracts_in_MC-Review.xlsx)")
    ap.add_argument("--database-url", default=os.environ.get("DATABASE_URL"),
                    help="postgres connection string (defaults to $DATABASE_URL)")
    args = ap.parse_args()

    if not args.database_url:
        sys.exit("No database URL. Set DATABASE_URL or pass --database-url.")
    if args.to_date and args.from_date and args.to_date < args.from_date:
        sys.exit(f"TO date {args.to_date} is before FROM date {args.from_date}.")
    out_path = args.out

    lookup = load_program_lookup()
    rows = run_query(args.database_url, args.from_date, args.to_date, args.by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["DSNP Contract Submissions"])
    ws.append(HEADERS)
    for r in rows:
        program_ids = r["program_ids"].split(",") if r["program_ids"] else []
        name = package_name(r["state_code"], int(r["state_number"]),
                            program_ids, lookup)
        ws.append([
            name,
            r["state_name"],
            r["contract_id"],
            r["federal_authorities"] or None,
            r["rate_medicaid_populations"] or None,
            r["dsnp_contract"] in ("t", "true", "True", "TRUE"),
            r["initial_submitted_date"] or None,
            r["latest_submitted_date"] or None,
        ])

    # Readability: title banner, bold + frozen header, content-width columns.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A3"
    # Width to content, ignoring the merged title row so it doesn't skew column A.
    for col in range(1, len(HEADERS) + 1):
        width = max(
            (len(str(ws.cell(row=r, column=col).value or ""))
             for r in range(2, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 80)

    wb.save(out_path)
    when = f"{args.by}-submitted"
    if args.from_date and args.to_date:
        window = f" {when} {args.from_date} to {args.to_date}"
    elif args.from_date:
        window = f" {when} on/after {args.from_date}"
    else:
        window = ""
    print(f"Wrote {len(rows)} DSNP contract(s){window} -> {out_path}")


if __name__ == "__main__":
    main()
